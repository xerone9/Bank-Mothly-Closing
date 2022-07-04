import os
from PyPDF2 import PdfFileMerger
from shutil import copyfile
import openpyxl as xl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
import datetime
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import time
from openpyxl.utils import get_column_letter


#
# ####################################### !!!!!!!!! CREATING FOLDERS !!!!!!!!!!!!!!!!!! ############################################################
#
Month = input("Enter Month Name: ")
print("")
print("Creating Folders...")
print("")

Years = Month[-4] + Month[-3] + Month[-2] + Month[-1]
Year = str(Years)
MonthSingle = Month[0] + Month[1] + Month[2]
monther = {
    "Feb": "Jan",
    "Mar": "Feb",
    "Apr": "Mar",
    "May": "Apr",
    "Jun": "May",
    "Jul": "Jun",
    "Aug": "Jul",
    "Sep": "Aug",
    "Oct": "Sep",
    "Nov": "Oct",
    "Dec": "Nov",
}
oldMonth = monther.get(MonthSingle, MonthSingle) + " " + Years
if MonthSingle == "Jan":
    Year = int(Years) - 1
    oldMonth = "Dec " + str(Year)

Month2021 = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month)
Bank336 = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month + "\\33-6")
Bank85 = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month + "\\8-5")
Bank74 = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month + "\\7-4")
Bank41 = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month + "\\4-1")
Bank63 = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month + "\\6-3")
Bank66 = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month + "\\6-6")
Bank100 = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month + "\\10-0")
Bank507 = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month + "\\50-7")
Bank55 = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month + "\\5-5")
try:
    os.mkdir(Month2021)
    print("Successfully created the directory %s" % Month2021)
    os.mkdir(Bank336)
    print ("Successfully created the directory %s " % Bank336)
    os.mkdir(Bank85)
    print ("Successfully created the directory %s " % Bank85)
    os.mkdir(Bank74)
    print ("Successfully created the directory %s " % Bank74)
    os.mkdir(Bank41)
    print ("Successfully created the directory %s " % Bank41)
    os.mkdir(Bank63)
    print ("Successfully created the directory %s " % Bank63)
    os.mkdir(Bank66)
    print ("Successfully created the directory %s " % Bank66)
    os.mkdir(Bank100)
    print ("Successfully created the directory %s " % Bank100)
    os.mkdir(Bank507)
    print ("Successfully created the directory %s " % Bank507)
    os.mkdir(Bank55)
    print ("Successfully created the directory %s " % Bank55)
except FileExistsError:
    input("Some Files/Folders Already Exist. If you had made any modification Close or module or Press Enter To Continue...")
    pass

# ####################################### !!!!!!!!! Copyting Files (Statements) !!!!!!!!!!!!!!!!!! ###################################################
#

print("")
print("Copying Account Statements To Respected Folders...")
print("")
desktop = os.path.expanduser("~\desktop\\Daily Work\\" +  Month + "\\")

for root, dirs, files in os.walk(desktop):
    for file in files:
        if file.startswith('0860380800055'):
            pather = desktop + Month + "\\5-5\\" + file
            if os.path.isfile(pather) and os.access(pather, os.R_OK):
                pass
            else:
                copyfile(root + "\\" + file,
                         Bank55 + "\\" + file)
                print(file)
        elif file.startswith('0860380800066'):
            pather = desktop + Month + "\\6-6\\" + file
            if os.path.isfile(pather) and os.access(pather, os.R_OK):
                pass
            else:
                copyfile(root + "\\" + file,
                         Bank66 + "\\" + file)
                print(file)
        elif file.startswith('0861480800041'):
            pather = desktop + Month + "\\4-1\\" + file
            if os.path.isfile(pather) and os.access(pather, os.R_OK):
                pass
            else:
                copyfile(root + "\\" + file,
                         Bank41 + "\\" + file)
                print(file)
        elif file.startswith('0861480800063'):
            pather = desktop + Month + "\\6-3\\" + file
            if os.path.isfile(pather) and os.access(pather, os.R_OK):
                pass
            else:
                copyfile(root + "\\" + file,
                         Bank63 + "\\" + file)
                print(file)
        elif file.startswith('0861480800074'):
            pather = desktop + Month + "\\7-4\\" + file
            if os.path.isfile(pather) and os.access(pather, os.R_OK):
                pass
            else:
                copyfile(root + "\\" + file,
                         Bank74 + "\\" + file)
                print(file)
        elif file.startswith('0861480800085'):
            pather = desktop + Month + "\\8-5\\" + file
            if os.path.isfile(pather) and os.access(pather, os.R_OK):
                pass
            else:
                copyfile(root + "\\" + file,
                         Bank85 + "\\" + file)
                print(file)
        elif file.startswith('0861480800100'):
            pather = desktop + Month + "\\10-0\\" + file
            if os.path.isfile(pather) and os.access(pather, os.R_OK):
                pass
            else:
                copyfile(root + "\\" + file,
                         Bank100 + "\\" + file)
                print(file)
        elif file.startswith('0861480800336'):
            pather = desktop + Month + "\\33-6\\" + file
            if os.path.isfile(pather) and os.access(pather, os.R_OK):
                pass
            else:
                copyfile(root + "\\" + file,
                         Bank336 + "\\" + file)
                print(file)
        elif file.startswith('0861480800507'):
            pather = desktop + Month + "\\50-7\\" + file
            if os.path.isfile(pather) and os.access(pather, os.R_OK):
                pass
            else:
                copyfile(root + "\\" + file,
                         Bank507 + "\\" + file)
                print(file)
#
# ####################################### !!!!!!!!! Merging PDF !!!!!!!!!!!!!!!!!! ############################################################
#
print("")
print("")
print("Make Sure that you had edited all the pdf bank statements (Corrections & Unknown Deposits). This Porcess Will Merge all PDF Files into one File For Printing")
Preparing_Merge_File = input("Press Enter For Merge File....")
print("")
print("")

desktop = os.path.expanduser(Month2021 + "\\")
file_dict = {}
for subdir, dirs, files in os.walk(desktop):
    for file in reversed(files):
        filepath = subdir + os.sep + file
        if file.startswith('0861480800336'):
            if filepath.endswith((".pdf", ".PDF")):
                file_dict[file] = filepath
for subdir, dirs, files in os.walk(desktop):
    for file in reversed(files):
        filepath = subdir + os.sep + file
        if file.startswith('0861480800063'):
            if filepath.endswith((".pdf", ".PDF")):
                file_dict[file] = filepath
for subdir, dirs, files in os.walk(desktop):
    for file in reversed(files):
        filepath = subdir + os.sep + file
        if file.startswith('0860380800055'):
             if filepath.endswith((".pdf", ".PDF")):
                file_dict[file] = filepath
for subdir, dirs, files in os.walk(desktop):
    for file in reversed(files):
        filepath = subdir + os.sep + file
        if file.startswith('0861480800041'):
            if filepath.endswith((".pdf", ".PDF")):
                file_dict[file] = filepath
for subdir, dirs, files in os.walk(desktop):
    for file in reversed(files):
        filepath = subdir + os.sep + file
        if file.startswith('0861480800085'):
            if filepath.endswith((".pdf", ".PDF")):
                    file_dict[file] = filepath
for subdir, dirs, files in os.walk(desktop):
    for file in reversed(files):
        filepath = subdir + os.sep + file
        if file.startswith('0861480800100'):
            if filepath.endswith((".pdf", ".PDF")):
                file_dict[file] = filepath
for subdir, dirs, files in os.walk(desktop):
    for file in reversed(files):
        filepath = subdir + os.sep + file
        if file.startswith('0861480800507'):
            if filepath.endswith((".pdf", ".PDF")):
                file_dict[file] = filepath
for subdir, dirs, files in os.walk(desktop):
    for file in reversed(files):
        filepath = subdir + os.sep + file
        if file.startswith('0861480800074'):
            if filepath.endswith((".pdf", ".PDF")):
                file_dict[file] = filepath
for subdir, dirs, files in os.walk(desktop):
    for file in reversed(files):
        filepath = subdir + os.sep + file
        if file.startswith('0860380800066'):
            if filepath.endswith((".pdf", ".PDF")):
                file_dict[file] = filepath

# use strict = False to ignore PdfReadError: Illegal character error
merger = PdfFileMerger(strict=False)

for k, v in file_dict.items():
    print(k, v)
    merger.append(v)

merger.write(desktop + Month + ".pdf")
location = desktop + Month + ".pdf"
print("File Saved Location " + location)


######################################## Excel File Creating ######################################################

root = os.path.expanduser("~\desktop\\" + Month + ".xlsx")
unroot = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month + "\\" + Month + ".xlsx")
print("")
print("""
Now Download the report from INDUS ERP --> Financeial Reports -> Set Month Stating and ending date and Select Account Wise Data
Once its done Upload file to online2pdf website and convert file to xlsx and place it to Desktop
""")
print("")
input("Make sure that converted File is on desktop having name '" + Month + ".xlsx'. Press Enter To Continue...")
copyfile(root, unroot)
print("")
print("Creating Excel File...")

desktop = os.path.expanduser("~\desktop\\Daily Work\\" + Month + "\\" + Month + "\\" + Month + ".xlsx")
wb = xl.load_workbook(desktop)
lastClosingFile = os.path.expanduser("~\desktop\\Daily Work\\" + oldMonth + "\\" + oldMonth + "\\" + oldMonth + ".xlsx")
if oldMonth == "Dec " + str(Year):
    lastClosingFile = os.path.expanduser(
        "~\desktop\\Daily Work\\" + str(Year) + "\\" + oldMonth + "\\" + oldMonth + "\\" + oldMonth + ".xlsx")
cb = xl.load_workbook(lastClosingFile)


print("Accessing Excel File...")
time.sleep(1.5)
sheet = wb['Table 1']
sheet2 = wb.create_sheet("Account Statements")
sheetOld = cb['Account Statements']
ws = wb['Account Statements']

print("")
print("UnMerging Cells (Will be Merged Later Onwards)...")
print("")

for items in sorted(sheet.merged_cells.ranges):
    sheet.unmerge_cells(str(items))
sheet.insert_cols(3, 1)

print("Removing Breakups Where One Account Contain Multiple Headers due to Page Break...")
time.sleep(1.5)


for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 1)
    if cell.value == "Sno":
        vell = sheet.cell(row + 1, 1)
        try:
            int(vell.value)
            if vell.value > 1:
                sheet.delete_rows(row - 1)
                sheet.delete_rows(row - 1)
        except TypeError:
            sheet.delete_rows(row - 1)
            sheet.delete_rows(row - 1)

print("")
print("Setting/Creating Columns [S.No, Transaction Date, Total No Of Deposits, Dr, Cr, Balance]...")
time.sleep(1.5)


for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 1)
    if cell.value == "Sno":
        sheet.insert_rows(row + 1)
        openingBalance = sheet.cell(row + 1, 2)
        openingBalance.value = "Opening Balance"
        sNo = sheet.cell(row + 1, 1)
        sNo.value = "0"
        totalDeposits = sheet.cell(row, 3)
        totalDeposits.value = "Total No. Of Deposits"
        transactionDate = sheet.cell(row, 2)
        transactionDate.value = "Transaction Date"
        dr = sheet.cell(row, 4)
        dr.value = "Dr"
        balance = sheet.cell(row, 6)
        balance.value = "Balance"

print("")
print("Adding Account Numbers To Account Names (Account Numbers are not in Sharafat Report)...")
time.sleep(1.5)


for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 1)
    if cell.value == "IU=PNBMDCF 10-0":
        cell.value = "IU=PNBMDCF 10-0 (0861480800100)"
        print(cell.value)

    if cell.value == "Indus Institute of Higher Education 4-1 {IIHE}":
        cell.value = "Indus Institute of Higher Education (0861480800041)"
        print(cell.value)
        time.sleep(1)
    if cell.value == "Indus University":
        cell.value = "Indus University (0861480800336)"
        print(cell.value)
        time.sleep(1)
    if cell.value == "Indus University 50-7 {Degree}":
        cell.value = "Indus University Degree (0861480800507)"
        print(cell.value)
        time.sleep(1)
    if cell.value == "Indus University 6-3 {Education}":
        cell.value = "Indus University Education (0861480800063)"
        print(cell.value)
        time.sleep(1)
    if cell.value == "Indus University 7-4 {SC}":
        cell.value = "Indus University SC (0861480800074)"
        print(cell.value)
        time.sleep(1)
    if cell.value == "Indus University {Technology}":
        cell.value = "Indus University Technology (0861480800085)"
        print(cell.value)
        time.sleep(1)
print("")

print("Rearranging Accounts Data with respect to our Bank Banks Order...")
time.sleep(1)


AllBank = []
Bank336 = []
for row in range(1, sheet.max_row):
    cell = sheet.cell(row, 1)
    if cell.value == "Indus University (0861480800336)":
        print("Book 1 - First Account - Indus University (0861480800336)")
        for drow in range(row, sheet.max_row):
            cell = sheet.cell(drow, 1)
            if cell.value is not None:
                Bank336.append(drow)
                AllBank.append(drow)
            else:
                break

time.sleep(1)

Bank63 = []
for row in range(1, sheet.max_row):
    cell = sheet.cell(row, 1)
    if cell.value == "Indus University Education (0861480800063)":
        print("Book 1 - Second Account - Indus University Education (0861480800063)")
        for drow in range(row, sheet.max_row):
            cell = sheet.cell(drow, 1)
            if cell.value is not None:
                Bank63.append(drow)
                AllBank.append(drow)
            else:
                break
    else:
        pass
time.sleep(1)

Bank41 = []
for row in range(1, sheet.max_row):
    cell = sheet.cell(row, 1)
    if cell.value == "Indus Institute of Higher Education (0861480800041)":
        print("Book 1 - Third Account " + cell.value)
        for drow in range(row, sheet.max_row):
            cell = sheet.cell(drow, 1)
            if cell.value is not None:
                Bank41.append(drow)
                AllBank.append(drow)
            else:
                break
    else:
        pass
time.sleep(1)
Bank85 = []
for row in range(1, sheet.max_row):
    cell = sheet.cell(row, 1)
    if cell.value == "Indus University Technology (0861480800085)":
        print("Book 2 - First Account - " + cell.value)
        for drow in range(row, sheet.max_row):
            cell = sheet.cell(drow, 1)
            if cell.value is not None:
                Bank85.append(drow)
                AllBank.append(drow)
            else:
                break
    else:
        pass
time.sleep(1)
Bank100 = []
for row in range(1, sheet.max_row):
    cell = sheet.cell(row, 1)
    if cell.value == "IU=PNBMDCF 10-0 (0861480800100)":
        print("Book 2 - Second Account - " + cell.value)
        for drow in range(row, sheet.max_row):
            cell = sheet.cell(drow, 1)
            if cell.value is not None:
                Bank100.append(drow)
                AllBank.append(drow)
            else:
                break
    else:
        pass
time.sleep(1)
Bank507 = []
for row in range(1, sheet.max_row):
    cell = sheet.cell(row, 1)
    if cell.value == "Indus University Degree (0861480800507)":
        print("Book 2 - Third Account - " + cell.value)
        for drow in range(row, sheet.max_row):
            cell = sheet.cell(drow, 1)
            if cell.value is not None:
                Bank507.append(drow)
                AllBank.append(drow)
            else:
                break
    else:
        pass
time.sleep(1)
Bank74 = []
for row in range(1, sheet.max_row):
    cell = sheet.cell(row, 1)
    if cell.value == "Indus University SC (0861480800074)":
        print("Book 2 - Forth Account - " + cell.value)
        for drow in range(row, sheet.max_row):
            cell = sheet.cell(drow, 1)
            if cell.value is not None:
                Bank74.append(drow)
                AllBank.append(drow)
            else:
                break
    else:
        pass

mr = sheet.max_row
mc = sheet.max_column
k = 1
l = 1
m = 1

# copying the cell values from source
# excel file to destination excel file

for i in (AllBank):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = sheet.cell(row=i, column=j)
        # writing the read value to destination excel file
        sheet2.cell(row=k, column=j).value = c.value
        l += 1
        if l < 7:
            k = m

        else:
            m += 1
            k = m
            l = 1

for row in range(1, sheet2.max_row):
    for col in range(1, sheet2.max_column):
        cell = sheet2.cell(row, col)
        cell.font = 'Arial'

for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "IU=PNBMDCF 10-0 (0861480800100)":
        grow = row
sheet2.insert_rows(grow)
for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "Indus Institute of Higher Education (0861480800041)":
        grow = row
sheet2.insert_rows(grow)
for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "Indus University Degree (0861480800507)":
        grow = row
sheet2.insert_rows(grow)
for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "Indus University Education (0861480800063)":
        grow = row
sheet2.insert_rows(grow)
for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "Indus University SC (0861480800074)":
        grow = row
sheet2.insert_rows(grow)
for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "Indus University Technology (0861480800085)":
        grow = row
sheet2.insert_rows(grow)

for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row, 1)
    if cell.value is None:
        cell.value = "Total: "
cell = sheet2.cell(sheet2.max_row + 1, 1)
cell.value = "Total: "

for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "IU=PNBMDCF 10-0 (0861480800100)":
        grow = row
sheet2.insert_rows(grow)
for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "Indus Institute of Higher Education (0861480800041)":
        grow = row
sheet2.insert_rows(grow)
for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "Indus University Degree (0861480800507)":
        grow = row
sheet2.insert_rows(grow)
for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "Indus University Education (0861480800063)":
        grow = row
sheet2.insert_rows(grow)
for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "Indus University SC (0861480800074)":
        grow = row
sheet2.insert_rows(grow)
for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "Indus University Technology (0861480800085)":
        grow = row

print("")
print("Adding Rows Below Book (To Sepearte it from Book 2)...")
time.sleep(1.5)


sheet2.insert_rows(grow)
sheet2.insert_rows(grow)
sheet2.insert_rows(grow)
sheet2.insert_rows(grow)
sheet2.insert_rows(grow)
sheet2.insert_rows(grow)

print("Merging Rows Containing Account Names...")
time.sleep(1.5)


mergeCell = []
for row in range(1, sheet2.max_row + 1):
    cell = sheet2.cell(row, 1)
    if cell.value == "IU=PNBMDCF 10-0 (0861480800100)":
        mergeCell.append(row)
    if cell.value == "Indus Institute of Higher Education (0861480800041)":
        mergeCell.append(row)
    if cell.value == "Indus University Degree (0861480800507)":
        mergeCell.append(row)
    if cell.value == "Indus University Education (0861480800063)":
        mergeCell.append(row)
    if cell.value == "Indus University SC (0861480800074)":
        mergeCell.append(row)
    if cell.value == "Indus University Technology (0861480800085)":
        mergeCell.append(row)
    if cell.value == "Indus University (0861480800336)":
        mergeCell.append(row)

total = []
for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row, 1)
    if cell.value == "Total: ":
        total.append(row)

al = Alignment(horizontal='center', vertical='center')
redFill = PatternFill(start_color='D6D6D6',
                      end_color='D6D6D6',
                      fill_type='solid')
for row in mergeCell:
    sheet2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = sheet2.cell(row, 1)
    cell.alignment = al
    cell.fill = redFill

    cell.font = Font(bold=True, name='Arial')

for row in total:
    sheet2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
sheet2.merge_cells(start_row=sheet2.max_row, start_column=1, end_row=sheet2.max_row, end_column=2)

cell = sheet2.cell(4, 2)

print("Refromatting Dates As Per Our Formatting...")
time.sleep(1.5)


for row in range(1, sheet2.max_row + 2):
    cell = sheet2.cell(row, 2)
    if len(str(cell.value)) == 19:
        correction = list(str(cell.value))
        date = correction[2] + correction[3] + correction[4] + correction[5] + correction[6] + correction[7] + \
               correction[0] + correction[1] + correction[8] + correction[9]
        cell.value = date


def listToStringWithoutBrackets(value):
    return str(value).replace('[', '').replace(']', '')


drange = 1
crow = 1
grandTotal = 0
totalLocation = []
rowsToBeAdded = []
rowsWiseTotal = []
updateRowsToBeAdded = []
updater = 0
debitTotal = 0

print("Formatting (Designing) Sheet...")
time.sleep(1.5)


for row in range(drange, sheet2.max_row):
    cell = sheet2.cell(row, 1)
    drange += 1
    # print(cell.value)
    if cell.value == "0":
        for drow in range(drange, sheet2.max_row):
            cell = sheet2.cell(drow, 1)
            if cell.value != "Total: ":
                rowsToBeAdded.append(drow)
            else:
                break

    elif cell.value == "Total: ":
        totalLocation.append(row)
        # print(row)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


for row in rowsToBeAdded:
    cell = sheet2.cell(row, 1)
    cell.border = thin_border
    cell = sheet2.cell(row, 2)
    cell.border = thin_border
    cell = sheet2.cell(row, 3)
    cell.border = thin_border
    cell = sheet2.cell(row, 4)
    cell.border = thin_border
    cell = sheet2.cell(row, 5)
    cell.border = thin_border
    cell = sheet2.cell(row, 6)
    cell.border = thin_border

al = Alignment(horizontal='center', vertical='center')
right = Alignment(horizontal='right', vertical='center')
left = Alignment(horizontal='left', vertical='center')
redFill = PatternFill(start_color='F5F5F5',
                      end_color='F5F5F5',
                      fill_type='solid')


# Selecting Rows and Pasting it in sheet 2. Used Try because the account 6-3 is dormant and not it is throwing index error. Working is accurate but now 6-3 array is taking the values of the next account. and same mechanics is below

try:
    total336 = []
    acid = rowsToBeAdded[0]
    for rows in rowsToBeAdded:
        acid += 1
        ggwp = acid - rows
        if ggwp > 0:
            total336.append(rows)
        else:
            pass

    for rows in total336:
        rowsToBeAdded.pop(0)


    total63 = []
    acid = rowsToBeAdded[0]
    for rows in rowsToBeAdded:
        acid += 1
        ggwp = acid - rows
        if ggwp > 0:
            total63.append(rows)
        else:
            pass

    for rows in total63:
        rowsToBeAdded.pop(0)


    total41 = []
    acid = rowsToBeAdded[0]
    for rows in rowsToBeAdded:
        acid += 1
        ggwp = acid - rows
        if ggwp > 0:
            total41.append(rows)
        else:
            pass

    for rows in total41:
        rowsToBeAdded.pop(0)


    total85 = []
    acid = rowsToBeAdded[0]
    for rows in rowsToBeAdded:
        acid += 1
        ggwp = acid - rows
        if ggwp > 0:
            total85.append(rows)
        else:
            pass

    for rows in total85:
        rowsToBeAdded.pop(0)


    total100 = []
    acid = rowsToBeAdded[0]
    for rows in rowsToBeAdded:
        acid += 1
        ggwp = acid - rows
        if ggwp > 0:
            total100.append(rows)
        else:
            pass

    for rows in total100:
        rowsToBeAdded.pop(0)


    total507 = []
    acid = rowsToBeAdded[0]
    for rows in rowsToBeAdded:
        acid += 1
        ggwp = acid - rows
        if ggwp > 0:
            total507.append(rows)
        else:
            pass


    for rows in total507:
        rowsToBeAdded.pop(0)


    total74 = []
    acid = rowsToBeAdded[0]
    for rows in rowsToBeAdded:
        acid += 1
        ggwp = acid - rows
        if ggwp > 0:
            total74.append(rows)
        else:
            pass
    for rows in total74:
        rowsToBeAdded.pop(0)

except IndexError:
    pass

print("Totalling Debit And Credit Columns...")
time.sleep(1.5)

for throw in total336:
    cell = sheet2.cell(throw, 4)
    if cell.value is None:
        pass
    else:
        try:
            grandTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            grandTotal += cell.value
            break
for throw in total336:
    cell = sheet2.cell(throw, 5)
    if cell.value is None:
        pass
    else:
        try:
            debitTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            debitTotal += cell.value
            break
total336 = grandTotal
grandTotal = 0
debit336 = debitTotal
debitTotal = 0

for throw in total63:
    cell = sheet2.cell(throw, 4)
    if cell.value is None:
        pass
    else:
        try:
            grandTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            grandTotal += cell.value
            break

for throw in total63:
    cell = sheet2.cell(throw, 5)
    if cell.value is None:
        pass
    else:
        try:
            debitTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            debitTotal += cell.value

total63 = grandTotal
grandTotal = 0
debit63 = debitTotal
debitTotal = 0

for throw in total41:
    cell = sheet2.cell(throw, 4)
    if cell.value is None:
        pass
    else:
        try:
            grandTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            grandTotal += cell.value

for throw in total41:
    cell = sheet2.cell(throw, 5)
    if cell.value is None:
        pass
    else:
        try:
            debitTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            debitTotal += cell.value

total41 = grandTotal
grandTotal = 0
debit41 = debitTotal
debitTotal = 0

for throw in total85:
    cell = sheet2.cell(throw, 4)
    if cell.value is None:
        pass
    else:
        try:
            grandTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            grandTotal += cell.value

for throw in total85:
    cell = sheet2.cell(throw, 5)
    if cell.value is None:
        pass
    else:
        try:
            debitTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            debitTotal += cell.value

total85 = grandTotal
grandTotal = 0
debit85 = debitTotal
debitTotal = 0

for throw in total100:
    cell = sheet2.cell(throw, 4)
    if cell.value is None:
        pass
    else:
        try:
            grandTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            grandTotal += cell.value

for throw in total100:
    cell = sheet2.cell(throw, 5)
    if cell.value is None:
        pass
    else:
        try:
            debitTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            debitTotal += cell.value


total100 = grandTotal
grandTotal = 0
debit100 = debitTotal
debitTotal = 0

for throw in total507:
    cell = sheet2.cell(throw, 4)
    if cell.value is None:
        pass
    else:
        try:
            grandTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            grandTotal += cell.value

for throw in total507:
    cell = sheet2.cell(throw, 5)
    if cell.value is None:
        pass
    else:
        try:
            debitTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            debitTotal += cell.value

total507 = grandTotal
grandTotal = 0
debit507 = debitTotal
debitTotal = 0

for throw in total74:
    cell = sheet2.cell(throw, 4)
    if cell.value is None:
        pass
    else:
        try:
            grandTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            grandTotal += cell.value

for throw in total74:
    cell = sheet2.cell(throw, 5)
    if cell.value is None:
        pass
    else:
        try:
            debitTotal += cell.value
        except TypeError:
            value = str(cell.value)
            value1 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%m-%d-%y")
            cell.value = int(value1.replace("-", ""))
            debitTotal += cell.value
total74 = grandTotal
grandTotal = 0
debit74 = debitTotal
debitTotal = 0

for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row, 1)
    if cell.value == "Total: ":
        cell = sheet2.cell(row, 4)
        cell.value = grandTotal
        grandTotal = 0
        totalLocation.append(row)
        # print(row)

creditTotal = []
creditTotal.append(total336)
creditTotal.append(total63)
creditTotal.append(total41)
creditTotal.append(total85)
creditTotal.append(total100)
creditTotal.append(total507)
creditTotal.append(total74)

debitTotal = []
debitTotal.append(debit336)
debitTotal.append(debit63)
debitTotal.append(debit41)
debitTotal.append(debit85)
debitTotal.append(debit100)
debitTotal.append(debit507)
debitTotal.append(debit74)

for f, b, d in zip(creditTotal, totalLocation, debitTotal):
    cell = sheet2.cell(b, 4)
    cell.value = f
    cell = sheet2.cell(b, 5)
    cell.value = d

print("Adding Commas To Numeric Values...")
time.sleep(1.5)


LastBalances=[]
for row in range(1, sheetOld.max_row):
    cell = sheetOld.cell(row, 1)
    if cell.value == "Indus University (0861480800336)":
        for throw in range(row, sheetOld.max_row):
            cell = sheetOld.cell(throw, 1)
            if cell.value == "Total: ":
                cell = sheetOld.cell(throw-1, 6)
                Closing336 = int(str(cell.value).replace(',', ''))
                break
    if cell.value == "Indus University Education (0861480800063)":
        for throw in range(row, sheetOld.max_row):
            cell = sheetOld.cell(throw, 1)
            if cell.value == "Total: ":
                cell = sheetOld.cell(throw - 1, 6)
                Closing63 = int(str(cell.value).replace(',', ''))
                break
    if cell.value == "Indus Institute of Higher Education (0861480800041)":
        for throw in range(row, sheetOld.max_row):
            cell = sheetOld.cell(throw, 1)
            if cell.value == "Total: ":
                cell = sheetOld.cell(throw - 1, 6)
                Closing41 = int(str(cell.value).replace(',', ''))
                break
    if cell.value == "Indus University Technology (0861480800085)":
        for throw in range(row, sheetOld.max_row):
            cell = sheetOld.cell(throw, 1)
            if cell.value == "Total: ":
                cell = sheetOld.cell(throw - 1, 6)
                Closing85 = int(str(cell.value).replace(',', ''))
                break
    if cell.value == "IU=PNBMDCF 10-0 (0861480800100)":
        for throw in range(row, sheetOld.max_row):
            cell = sheetOld.cell(throw, 1)
            if cell.value == "Total: ":
                cell = sheetOld.cell(throw - 1, 6)
                Closing100 = int(str(cell.value).replace(',', ''))
                break
    if cell.value == "Indus University Degree (0861480800507)":
        for throw in range(row, sheetOld.max_row):
            cell = sheetOld.cell(throw, 1)
            if cell.value == "Total: ":
                cell = sheetOld.cell(throw - 1, 6)
                Closing507 = int(str(cell.value).replace(',', ''))
                break

cell = sheetOld.cell(sheetOld.max_row - 1, 6)
Closing74 = int(str(cell.value).replace(',', ''))

print("Fetching Closing Values From Last Month Closing File...")
time.sleep(1.5)


for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row, 1)
    if cell.value == "Indus University (0861480800336)":
        for throw in range(row, sheet2.max_row):
            cell = sheet2.cell(throw, 6)
            if cell.value == "Balance":
                cell = sheet2.cell(throw + 1, 6)
                cell.value = Closing336
                break
    if cell.value == "Indus University Education (0861480800063)":
        for throw in range(row, sheet2.max_row):
            cell = sheet2.cell(throw, 6)
            if cell.value == "Balance":
                cell = sheet2.cell(throw + 1, 6)
                cell.value = Closing63
                break
    if cell.value == "Indus Institute of Higher Education (0861480800041)":
        for throw in range(row, sheet2.max_row):
            cell = sheet2.cell(throw, 6)
            if cell.value == "Balance":
                cell = sheet2.cell(throw + 1, 6)
                cell.value = Closing41
                break
    if cell.value == "Indus University Technology (0861480800085)":
        for throw in range(row, sheet2.max_row):
            cell = sheet2.cell(throw, 6)
            if cell.value == "Balance":
                cell = sheet2.cell(throw + 1, 6)
                cell.value = Closing85
                break
    if cell.value == "IU=PNBMDCF 10-0 (0861480800100)":
        for throw in range(row, sheet2.max_row):
            cell = sheet2.cell(throw, 6)
            if cell.value == "Balance":
                cell = sheet2.cell(throw + 1, 6)
                cell.value = Closing100
                break
    if cell.value == "Indus University Degree (0861480800507)":
        for throw in range(row, sheet2.max_row):
            cell = sheet2.cell(throw, 6)
            if cell.value == "Balance":
                cell = sheet2.cell(throw + 1, 6)
                cell.value = Closing507
                break
    if cell.value == "Indus University SC (0861480800074)":
        for throw in range(row, sheet2.max_row):
            cell = sheet2.cell(throw, 6)
            if cell.value == "Balance":
                cell = sheet2.cell(throw + 1, 6)
                cell.value = Closing74
                break

print("Now Setting Values For Balance Column...")
time.sleep(1.5)


for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row, 6)
    if cell.value == "Balance":
        for throw in range(row, sheet2.max_row):
            cellBalance = sheet2.cell(throw + 1, 6)
            cellDebit = sheet2.cell(throw + 2, 4)
            cellCredit = sheet2.cell(throw + 2, 5)
            cellResult = sheet2.cell(throw + 2, 6)
            cellBreak = sheet2.cell(throw+3, 1)
            if cellBalance.value is None:
                cellBalance.value = 0
            if cellDebit.value is None:
                cellDebit.value = 0
            if cellCredit.value is None:
                cellCredit.value = 0
            if cellResult.value is None:
                cellResult.value = 0
            cellResult.value = cellBalance.value + cellDebit.value - cellCredit.value
            if cellBreak.value == "Total: ":
                break

for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row,6)
    cell.font = Font(name='Arial')

print("Formatting (Setting Arial Font) Sheet...")
time.sleep(1.5)


for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row, 4)
    try:
        number_with_commas = "{:,}".format(cell.value)
        cell.value = number_with_commas
        cell.alignment = right
    except TypeError:
        pass
    except ValueError:
        pass

for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row, 5)
    try:
        number_with_commas = "{:,}".format(cell.value)
        cell.value = number_with_commas
        cell.alignment = right
    except TypeError:
        pass
    except ValueError:
        pass

for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row, 6)
    try:
        number_with_commas = "{:,}".format(cell.value)
        cell.value = number_with_commas
        cell.alignment = right
    except TypeError:
        pass
    except ValueError:
        pass

for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row, 1)
    if cell.value == "Sno":
        cell.alignment = al
        cell.font = Font(bold=True, name='Arial')
        cell.border = thin_border
        cell.fill = redFill
        cell = sheet2.cell(row, 2)
        cell.font = Font(bold=True, name='Arial')
        cell.alignment = left
        cell.fill = redFill
        cell.border = thin_border
        cell = sheet2.cell(row, 3)
        cell.font = Font(bold=True, name='Arial')
        cell.alignment = right
        cell.fill = redFill
        cell.border = thin_border
        cell = sheet2.cell(row, 4)
        cell.font = Font(bold=True, name='Arial')
        cell.alignment = right
        cell.fill = redFill
        cell.border = thin_border
        cell = sheet2.cell(row, 5)
        cell.font = Font(bold=True, name='Arial')
        cell.alignment = right
        cell.fill = redFill
        cell.border = thin_border
        cell = sheet2.cell(row, 6)
        cell.font = Font(bold=True, name='Arial')
        cell.alignment = right
        cell.fill = redFill
        cell.border = thin_border
    if cell.value == "0":
        cell.border = thin_border
        cell.alignment = right
        cell = sheet2.cell(row, 2)
        cell.font = Font(bold=True, name='Arial')
        cell = sheet2.cell(row, 6)
        cell.font = Font(bold=True, name='Arial')
        cell.alignment = left
        cell.border = thin_border
        cell = sheet2.cell(row, 3)
        cell.alignment = right
        cell.border = thin_border
        cell = sheet2.cell(row, 4)
        cell.alignment = right
        cell.border = thin_border
        cell = sheet2.cell(row, 5)
        cell.alignment = right
        cell.border = thin_border
        cell = sheet2.cell(row, 6)
        cell.alignment = right
        cell.border = thin_border
    if cell.value == "Total: ":
        cell = sheet2.cell(row, 1)
        cell.font = Font(bold=True, name='Arial')
        cell.alignment = right
        cell = sheet2.cell(row, 3)
        cell.font = Font(bold=True, name='Arial')
        cell.alignment = right
        cell.fill = redFill
        cell.border = thin_border
        cell = sheet2.cell(row, 4)
        cell.font = Font(bold=True, name='Arial')
        cell.alignment = right
        cell.fill = redFill
        cell.border = thin_border
        cell = sheet2.cell(row-1, 6)
        cell.font = Font(bold=True, name='Arial')
        cell = sheet2.cell(row, 5)
        cell.font = Font(bold=True, name='Arial')
        cell.alignment = right
        cell.fill = redFill
        cell.border = thin_border

for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row, 1)
    if cell.value == "IU=PNBMDCF 10-0 (0861480800100)":
        cell.font = Font(bold=True, name='Arial')
    if cell.value == "Indus Institute of Higher Education (0861480800041)":
        cell.font = Font(bold=True, name='Arial')
    if cell.value == "Indus University Degree (0861480800507)":
        cell.font = Font(bold=True, name='Arial')
    if cell.value == "Indus University Education (0861480800063)":
        cell.font = Font(bold=True, name='Arial')
    if cell.value == "Indus University SC (0861480800074)":
        cell.font = Font(bold=True, name='Arial')
    if cell.value == "Indus University Technology (0861480800085)":
        cell.font = Font(bold=True, name='Arial')
    if cell.value == "Indus University (0861480800336)":
        cell.font = Font(bold=True, name='Arial')

print("Replacing Zeros With (Null) Blank Values...")
time.sleep(1.5)


for row in range(1, sheet2.max_row):
    cell = sheet2.cell(row, 4)
    if cell.value == "0":
        cell.value = None
    cell = sheet2.cell(row, 5)
    if cell.value == "0":
        cell.value = None

for row in range(1, sheet2.max_row):
    for col in range(1, sheet2.max_column):
        cell = sheet2.cell(row, col)
        try:
            int(cell.value)
        except ValueError:
            pass
        except TypeError:
            pass



del wb['Table 1']
wb.save(desktop)

print("Auto Fit Columns...")
time.sleep(1.5)


dim_holder = DimensionHolder(worksheet=ws)

for col in range(ws.min_column+1, ws.max_column + 1):
    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=13)

ws.column_dimensions = dim_holder
wb.save(desktop)

print("Completed...")
os.startfile(desktop)

time.sleep(1.5)
